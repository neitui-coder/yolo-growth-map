#!/usr/bin/env python3
"""Generate five 2026 YOLO+ member poster style samples from the new workbook."""

import io
import json
import math
import re
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFilter, ImageFont, ImageOps


REPO = Path(__file__).resolve().parents[1]
WORKBOOK = Path("/Users/Sean/Downloads/YOLO+2026年会员成长档案资料收集.xlsx")
OLD_POSTER_DIR = Path("/Users/Sean/Downloads/2024-2025年会员海报")
AVATAR_DIR = REPO / "assets-source" / "images" / "avatars"
MASTER_JSON = REPO / "data" / "yolo-2025-members.master.json"
OUTPUT_DIR = REPO / "outputs" / "poster-samples-2026"

W, H = 1587, 2245
MARGIN = 92

FONT_HEI = "/System/Library/Fonts/Hiragino Sans GB.ttc"
FONT_ST = "/System/Library/Fonts/STHeiti Medium.ttc"
FONT_DIN = "/System/Library/Fonts/Supplemental/DIN Condensed Bold.ttf"
FONT_GEORGIA_BOLD = "/System/Library/Fonts/Supplemental/Georgia Bold.ttf"


SELECTED = [
    {"name": "蒋柳璟子", "style": "classic", "avatar": "pdf2025-jingzi.jpg", "oldPoster": "蒋柳璟子-常务理事.jpg", "label": "A 复古霓虹"},
    {"name": "罗琦玥", "style": "editorial", "avatar": "pdf2025-sarah.jpg", "oldPoster": "罗琦玥-常务理事.jpg", "label": "B 杂志留白"},
    {"name": "王骁", "style": "tech", "avatar": "pdf2025-sean.jpg", "oldPoster": "王骁-常务理事.jpg", "label": "C 深蓝科技"},
    {"name": "吴昌鸿", "style": "riso", "avatar": "pdf2025-peter.jpg", "oldPoster": "吴昌鸿-副组长.jpg", "label": "D 活力拼贴"},
    {"name": "徐志谦", "style": "minimal", "avatar": "pdf2025-zhiqian.jpg", "oldPoster": "徐志谦-常务理事.jpg", "label": "E 极简档案"},
]


def font(size, bold=False, latin=False):
    path = FONT_DIN if latin else (FONT_ST if bold else FONT_HEI)
    return ImageFont.truetype(path, size)


def clean_name(value):
    return re.sub(r"\s+", "", str(value or "")).strip()


def normalize_date(value):
    if not value:
        return ""
    text = str(value)
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if m:
        return f"{int(m.group(2))}月{int(m.group(3))}日"
    return text.replace("00:00:00", "").strip()


def normalize_location(value):
    text = str(value or "").strip()
    if not text or text == "/":
        return ""
    text = text.replace("，", "-").replace(",", "-")
    text = re.sub(r"\s+", "", text)
    if text.startswith("中国-"):
        text = text[3:]
    return text.strip("-")


def normalize_slogan(value):
    text = str(value or "").strip()
    text = re.sub(r"^(slogan|Slogan|SLOGAN)\s*[：:]\s*", "", text)
    if re.sub(r"[\s,，.。]+", "", text).lower() == "lcamelsawlconquered":
        return "I came, I saw, I conquered"
    return text


def normalize_hobbies(value):
    text = str(value or "").strip()
    if not text or text == "/":
        return ""
    text = text.replace("，", "、").replace(",", "、").replace("/", "、")
    return re.sub(r"\s+", "", text)


def read_rows():
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb.active
    rows = {}
    for r in range(3, ws.max_row + 1):
        name = clean_name(ws.cell(r, 3).value)
        if not name:
            continue
        rows[name] = {
            "row": r,
            "category": str(ws.cell(r, 2).value or "").strip(),
            "name": name,
            "birthday": normalize_date(ws.cell(r, 7).value),
            "education": str(ws.cell(r, 8).value or "").strip(),
            "company": str(ws.cell(r, 9).value or "").strip(),
            "position": str(ws.cell(r, 10).value or "").strip(),
            "location": normalize_location(ws.cell(r, 11).value),
            "hobbies": normalize_hobbies(ws.cell(r, 14).value),
            "slogan": normalize_slogan(ws.cell(r, 15).value),
            "notes": str(ws.cell(r, 17).value or "").strip(),
        }

    embedded = {}
    for img in getattr(ws, "_images", []):
        anchor = img.anchor._from
        row = anchor.row + 1
        col = anchor.col + 1
        if col not in (13, 14):
            continue
        name = clean_name(ws.cell(row, 3).value)
        if not name:
            continue
        try:
            raw = img._data()
            image = Image.open(io.BytesIO(raw)).convert("RGB")
            embedded.setdefault(name, image)
        except Exception:
            pass
    if MASTER_JSON.exists():
        master = json.loads(MASTER_JSON.read_text(encoding="utf-8"))
        for record in master.get("members", []):
            name = clean_name(record.get("identity", {}).get("name"))
            if name not in rows:
                continue
            profile = record.get("profile", {})
            interests = record.get("interests", {})
            community = record.get("community", {})
            rows[name]["englishName"] = record.get("identity", {}).get("englishName") or ""
            rows[name]["hobbies"] = rows[name]["hobbies"] or normalize_hobbies("、".join(interests.get("hobbies") or []))
            rows[name]["gallup"] = " | ".join(profile.get("gallup") or [])
            if not rows[name]["category"] and community.get("role"):
                rows[name]["category"] = community.get("role")
    return rows, embedded


def avatar_for(member, embedded, item):
    if member["name"] in embedded:
        return embedded[member["name"]].copy()
    path = AVATAR_DIR / item["avatar"]
    if path.exists():
        return Image.open(path).convert("RGB")
    old = OLD_POSTER_DIR / item["oldPoster"]
    if old.exists():
        base = Image.open(old).convert("RGB")
        return base.crop((160, 520, 980, 1480))
    return Image.new("RGB", (900, 900), "#dbeafe")


def cover_crop(img, size, pos=(0.5, 0.5)):
    return ImageOps.fit(img, size, method=Image.Resampling.LANCZOS, centering=pos)


def round_rect(draw, xy, radius, fill, outline=None, width=1):
    draw.rounded_rectangle(xy, radius=radius, fill=fill, outline=outline, width=width)


def draw_text(draw, xy, text, fnt, fill, anchor=None, spacing=8, stroke_width=0, stroke_fill=None):
    draw.text(xy, text, font=fnt, fill=fill, anchor=anchor, spacing=spacing, stroke_width=stroke_width, stroke_fill=stroke_fill)


def text_width(draw, text, fnt):
    if not text:
        return 0
    return draw.textlength(text, font=fnt)


def wrap_text(draw, text, fnt, max_width, max_lines=None):
    text = str(text or "").replace("\n", " ").strip()
    if not text:
        return []
    units = []
    token = ""
    for ch in text:
        if ch.isspace():
            if token:
                units.append(token)
                token = ""
            if units and units[-1] != " ":
                units.append(" ")
            continue
        if ord(ch) < 128 and (ch.isalnum() or ch in ".+-/&"):
            token += ch
        else:
            if token:
                units.append(token)
                token = ""
            units.append(ch)
    if token:
        units.append(token)
    lines, line = [], ""
    for unit in units:
        if unit == " ":
            if line and not line.endswith(" "):
                line += " "
            continue
        trial = line + unit
        if line and text_width(draw, trial, fnt) > max_width:
            lines.append(line.rstrip())
            line = unit
            if max_lines and len(lines) >= max_lines:
                break
        else:
            line = trial
    if line and (not max_lines or len(lines) < max_lines):
        lines.append(line.rstrip())
    if max_lines and len(lines) == max_lines and len("".join(lines)) < len(text.replace(" ", "")):
        while text_width(draw, lines[-1] + "…", fnt) > max_width and len(lines[-1]) > 1:
            lines[-1] = lines[-1][:-1]
        lines[-1] += "…"
    return lines


def draw_wrapped(draw, xy, text, fnt, fill, max_width, line_height, max_lines=None, anchor=None):
    lines = wrap_text(draw, text, fnt, max_width, max_lines)
    x, y = xy
    for i, line in enumerate(lines):
        draw.text((x, y + i * line_height), line, font=fnt, fill=fill, anchor=anchor)
    return y + len(lines) * line_height


def paste_rounded(canvas, img, xy, size, radius=48, border=None, border_width=0):
    photo = cover_crop(img, size, pos=(0.5, 0.38)).convert("RGBA")
    mask = Image.new("L", size, 0)
    md = ImageDraw.Draw(mask)
    md.rounded_rectangle((0, 0, size[0], size[1]), radius=radius, fill=255)
    canvas.paste(photo, xy, mask)
    if border:
        d = ImageDraw.Draw(canvas)
        d.rounded_rectangle((xy[0], xy[1], xy[0] + size[0], xy[1] + size[1]), radius=radius, outline=border, width=border_width)


def paste_circle(canvas, img, center, diameter, border="#fff", border_width=16, glow=None):
    if glow:
        glow_layer = Image.new("RGBA", (W, H), (0, 0, 0, 0))
        gd = ImageDraw.Draw(glow_layer)
        gd.ellipse((center[0] - diameter // 2 - 12, center[1] - diameter // 2 - 12, center[0] + diameter // 2 + 12, center[1] + diameter // 2 + 12), fill=glow)
        glow_layer = glow_layer.filter(ImageFilter.GaussianBlur(18))
        canvas.alpha_composite(glow_layer)
    photo = cover_crop(img, (diameter, diameter), pos=(0.5, 0.32)).convert("RGBA")
    mask = Image.new("L", (diameter, diameter), 0)
    md = ImageDraw.Draw(mask)
    md.ellipse((0, 0, diameter, diameter), fill=255)
    xy = (center[0] - diameter // 2, center[1] - diameter // 2)
    canvas.paste(photo, xy, mask)
    d = ImageDraw.Draw(canvas)
    d.ellipse((xy[0], xy[1], xy[0] + diameter, xy[1] + diameter), outline=border, width=border_width)


def info_lines(member):
    company = member["company"] if member["company"] != "/" else ""
    position = member["position"] if member["position"] != "/" else ""
    work = " / ".join([x for x in [company, position] if x])
    return [
        ("生日", member["birthday"]),
        ("现居", member["location"]),
        ("毕业院校", member["education"]),
        ("公司职位", work),
        ("兴趣爱好", member["hobbies"]),
        ("盖洛普TOP5", member.get("gallup", "")),
    ]


def draw_info_panel(draw, member, x, y, width, palette, compact=False):
    label_font = font(34 if not compact else 29, bold=True)
    body_font = font(39 if not compact else 32, bold=True)
    line_h = 55 if not compact else 47
    for label, value in info_lines(member):
        if not value:
            continue
        draw.text((x, y), f"{label}：", font=label_font, fill=palette["label"])
        y = draw_wrapped(draw, (x + (150 if not compact else 128), y - 3), value, body_font, palette["text"], width - (150 if not compact else 128), line_h, max_lines=2)
        y += 22 if not compact else 15
    return y


def background_gradient(colors):
    img = Image.new("RGBA", (W, H), colors[0])
    pix = img.load()
    stops = [tuple(int(c[i:i + 2], 16) for i in (1, 3, 5)) for c in colors]
    for y in range(H):
        t = y / (H - 1)
        pos = t * (len(stops) - 1)
        idx = min(int(pos), len(stops) - 2)
        local = pos - idx
        c = tuple(int(stops[idx][i] * (1 - local) + stops[idx + 1][i] * local) for i in range(3))
        for x in range(W):
            pix[x, y] = c + (255,)
    return img


def draw_logo(draw, x, y, dark=False):
    draw.text((x, y), "YOLO", font=font(82, bold=True, latin=True), fill="#1d4ed8" if not dark else "#ffffff")
    draw.text((x + 196, y + 3), "+", font=font(72, bold=True, latin=True), fill="#f97316" if not dark else "#38bdf8")


def draw_border_words(canvas, word, color, accent):
    d = ImageDraw.Draw(canvas)
    d.rectangle((0, 0, 96, H), fill=accent)
    d.rectangle((W - 96, 0, W, H), fill="#fde047")
    d.polygon([(0, H - 260), (W, H - 400), (W, H - 315), (0, H - 170)], fill="#fb7185")
    text = (" " + word) * 12
    d.text((120, 32), text, font=font(42, bold=True, latin=True), fill=color)
    d.text((150, H - 250), text, font=font(42, bold=True, latin=True), fill="#ffffff")
    side = Image.new("RGBA", (H, 84), (0, 0, 0, 0))
    sd = ImageDraw.Draw(side)
    sd.text((0, 12), text, font=font(38, bold=True, latin=True), fill=(255, 255, 255, 150))
    canvas.alpha_composite(side.rotate(90, expand=True), (-5, 70))
    canvas.alpha_composite(side.rotate(90, expand=True), (W - 92, 90))


def style_classic(member, photo):
    canvas = background_gradient(["#19c2d3", "#fff6a5", "#ffffff"])
    d = ImageDraw.Draw(canvas)
    draw_border_words(canvas, "GRATITUDE EXCELLENCE GROWTH", (255, 255, 255, 130), "#14b8c8")
    round_rect(d, (104, 118, W - 104, 1908), 58, (255, 255, 255, 188), outline=(255, 255, 255, 230), width=5)
    draw_logo(d, W - 500, 125)
    d.text((150, 165), "感恩 / 卓越 / 成长", font=font(60, bold=True), fill="#ffffff")
    d.text((150, 300), "万物皆可期", font=font(88, bold=True), fill="#a855f7", stroke_width=3, stroke_fill="#22d3ee")
    d.text((150, 438), member["name"], font=font(118, bold=True), fill="#f97316", stroke_width=3, stroke_fill="#2dd4bf")
    paste_circle(canvas, photo, (510, 815), 520, border="#fbbf24", border_width=20, glow=(251, 191, 36, 100))
    round_rect(d, (980, 690, 1378, 765), 10, "#a7f3d0")
    d.text((1035, 706), member["category"], font=font(38, bold=True), fill="#be185d")
    round_rect(d, (1045, 910, 1370, 980), 10, "#ccfbf1")
    d.text((1085, 926), "现居：" + (member["location"] or "-"), font=font(35, bold=True), fill="#1d4ed8")
    d.text((170, 1198), "“" + (member["slogan"] or "持续生长，保持好奇") + "”", font=font(50, bold=True), fill="#0f172a")
    draw_info_panel(d, member, 178, 1390, 1240, {"label": "#111827", "text": "#000000"}, compact=False)
    d.text((W - 138, H - 88), member["label"], font=font(30, bold=True), fill="#ffffff", anchor="ra")
    return canvas


def style_editorial(member, photo):
    canvas = Image.new("RGBA", (W, H), "#f7f0e8")
    d = ImageDraw.Draw(canvas)
    d.rectangle((0, 0, W, 275), fill="#0f172a")
    d.text((92, 72), "YOLO+ MEMBER ARCHIVE", font=font(42, bold=True, latin=True), fill="#f7f0e8")
    d.text((92, 138), "2026", font=font(90, bold=True, latin=True), fill="#fbbf24")
    draw_logo(d, W - 420, 86, dark=True)
    paste_rounded(canvas, photo, (835, 350), (540, 760), radius=40, border="#0f172a", border_width=6)
    d.text((102, 380), member["name"], font=font(132, bold=True), fill="#0f172a")
    d.text((108, 525), member["category"], font=font(42, bold=True), fill="#2563eb")
    d.line((104, 600, 770, 600), fill="#0f172a", width=7)
    draw_wrapped(d, (108, 665), member["slogan"] or "Every step counts.", font(62, bold=True), "#0f172a", 650, 76, max_lines=4)
    round_rect(d, (90, 1210, W - 90, 2058), 34, "#ffffff", outline="#0f172a", width=4)
    d.text((132, 1264), "PROFILE", font=font(50, bold=True, latin=True), fill="#2563eb")
    draw_info_panel(d, member, 132, 1360, 1240, {"label": "#2563eb", "text": "#0f172a"}, compact=True)
    d.text((W - 116, H - 92), member["label"], font=font(30, bold=True), fill="#0f172a", anchor="ra")
    return canvas


def style_tech(member, photo):
    canvas = Image.new("RGBA", (W, H), "#07111f")
    d = ImageDraw.Draw(canvas)
    for x in range(0, W, 86):
        d.line((x, 0, x, H), fill=(59, 130, 246, 30), width=1)
    for y in range(0, H, 86):
        d.line((0, y, W, y), fill=(59, 130, 246, 30), width=1)
    d.rectangle((0, 0, W, 330), fill="#0b1d34")
    d.text((92, 84), "GROWTH MAP", font=font(84, bold=True, latin=True), fill="#38bdf8")
    d.text((95, 180), "YOLO+ 2026 MEMBER PROFILE", font=font(38, bold=True, latin=True), fill="#93c5fd")
    draw_logo(d, W - 420, 100, dark=True)
    paste_circle(canvas, photo, (W - 380, 590), 430, border="#38bdf8", border_width=12, glow=(56, 189, 248, 110))
    d.text((92, 430), member["name"], font=font(132, bold=True), fill="#ffffff")
    d.text((98, 570), member["category"], font=font(44, bold=True), fill="#facc15")
    chips = [member["birthday"], member["location"], member["position"] if member["position"] != "/" else ""]
    cx, cy = 96, 675
    for chip in [c for c in chips if c]:
        tw = int(text_width(d, chip, font(33, bold=True))) + 44
        round_rect(d, (cx, cy, cx + tw, cy + 62), 31, "#12365c", outline="#2563eb", width=2)
        d.text((cx + 22, cy + 12), chip, font=font(33, bold=True), fill="#dbeafe")
        cx += tw + 18
    round_rect(d, (90, 860, W - 90, 2055), 38, (15, 23, 42, 220), outline="#1e40af", width=4)
    d.text((140, 922), "SIGNAL", font=font(48, bold=True, latin=True), fill="#38bdf8")
    draw_wrapped(d, (140, 1002), member["slogan"] or "Build with patience.", font(56, bold=True), "#ffffff", 1260, 70, max_lines=3)
    d.line((140, 1240, W - 140, 1240), fill="#1e40af", width=3)
    draw_info_panel(d, member, 140, 1312, 1220, {"label": "#60a5fa", "text": "#e5f3ff"}, compact=True)
    d.text((W - 112, H - 90), member["label"], font=font(30, bold=True), fill="#60a5fa", anchor="ra")
    return canvas


def style_riso(member, photo):
    canvas = Image.new("RGBA", (W, H), "#fff7ed")
    d = ImageDraw.Draw(canvas)
    d.polygon([(0, 0), (W, 0), (W, 620), (0, 360)], fill="#fb923c")
    d.polygon([(0, 1260), (W, 900), (W, H), (0, H)], fill="#14b8a6")
    d.ellipse((-160, 260, 620, 1040), fill="#fde047")
    d.ellipse((930, 260, 1680, 1020), fill="#f9a8d4")
    draw_logo(d, 102, 90)
    d.text((102, 210), "感恩  卓越  成长", font=font(52, bold=True), fill="#ffffff")
    paste_rounded(canvas, photo, (810, 340), (560, 780), radius=70, border="#111827", border_width=7)
    d.text((96, 470), member["name"], font=font(126, bold=True), fill="#111827", stroke_width=3, stroke_fill="#fde047")
    d.text((104, 605), member["category"], font=font(44, bold=True), fill="#ffffff")
    draw_wrapped(d, (104, 715), member["slogan"] or "Stay curious.", font(58, bold=True), "#111827", 660, 72, max_lines=4)
    round_rect(d, (92, 1260, W - 92, 2070), 48, "#ffffff", outline="#111827", width=5)
    draw_info_panel(d, member, 145, 1325, 1210, {"label": "#ea580c", "text": "#111827"}, compact=True)
    for i in range(9):
        d.line((116 + i * 150, 2130, 176 + i * 150, 2070), fill="#ffffff", width=9)
    d.text((W - 112, H - 90), member["label"], font=font(30, bold=True), fill="#ffffff", anchor="ra")
    return canvas


def style_minimal(member, photo):
    canvas = Image.new("RGBA", (W, H), "#ffffff")
    d = ImageDraw.Draw(canvas)
    d.rectangle((0, 0, 220, H), fill="#0f172a")
    d.text((72, 1620), "YOLO+ 2026", font=font(48, bold=True, latin=True), fill="#ffffff")
    d.text((288, 120), member["label"], font=font(34, bold=True), fill="#2563eb")
    d.text((288, 190), member["name"], font=font(142, bold=True), fill="#0f172a")
    d.line((288, 375, W - 110, 375), fill="#0f172a", width=6)
    paste_rounded(canvas, photo, (920, 470), (430, 560), radius=18, border="#0f172a", border_width=5)
    d.text((288, 470), member["category"], font=font(48, bold=True), fill="#2563eb")
    draw_wrapped(d, (288, 555), "“" + (member["slogan"] or "Growth is a practice.") + "”", font(62, bold=True), "#0f172a", 560, 78, max_lines=5)
    y = 1130
    for label, value in info_lines(member):
        if not value:
            continue
        label_text = "GALLUP TOP5" if label == "盖洛普TOP5" else label
        d.text((288, y), label_text, font=font(34, bold=True, latin=label_text.isascii()), fill="#94a3b8")
        y = draw_wrapped(d, (288, y + 52), value, font(43, bold=True), "#0f172a", 1010, 60, max_lines=3) + 44
        d.line((288, y - 16, W - 110, y - 16), fill="#e2e8f0", width=2)
    d.text((W - 112, H - 90), "GRATITUDE / EXCELLENCE / GROWTH", font=font(30, bold=True, latin=True), fill="#0f172a", anchor="ra")
    return canvas


STYLE_BUILDERS = {
    "classic": style_classic,
    "editorial": style_editorial,
    "tech": style_tech,
    "riso": style_riso,
    "minimal": style_minimal,
}


def create_contact_sheet(paths):
    thumbs = []
    for p in paths:
        im = Image.open(p).convert("RGB")
        im.thumbnail((330, 467), Image.Resampling.LANCZOS)
        thumbs.append((p, im.copy()))
    sheet = Image.new("RGB", (1900, 680), "#f8fafc")
    d = ImageDraw.Draw(sheet)
    d.text((56, 42), "YOLO+ 2026 会员海报风格样张", font=font(54, bold=True), fill="#0f172a")
    d.text((58, 110), "五种方向：复古霓虹 / 杂志留白 / 深蓝科技 / 活力拼贴 / 极简档案", font=font(32, bold=True), fill="#475569")
    x = 58
    for p, im in thumbs:
        sheet.paste(im, (x, 170))
        d.rounded_rectangle((x, 170, x + im.width, 170 + im.height), radius=8, outline="#cbd5e1", width=3)
        d.text((x, 650), p.stem.replace("poster-", ""), font=font(24, bold=True), fill="#0f172a")
        x += 365
    out = OUTPUT_DIR / "poster-style-contact-sheet.jpg"
    sheet.save(out, quality=94)
    return out


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows, embedded = read_rows()
    output_paths = []
    for item in SELECTED:
        member = rows[item["name"]]
        member["label"] = item["label"]
        photo = avatar_for(member, embedded, item)
        poster = STYLE_BUILDERS[item["style"]](member, photo).convert("RGB")
        filename = f"poster-{item['label'].split()[0]}-{member['name']}.jpg"
        out = OUTPUT_DIR / filename
        poster.save(out, quality=94)
        output_paths.append(out)
        print(out)
    print(create_contact_sheet(output_paths))


if __name__ == "__main__":
    main()
